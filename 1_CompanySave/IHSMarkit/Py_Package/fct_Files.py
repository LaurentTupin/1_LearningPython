try:
    import os, time
    import pandas as pd
    import datetime as dt
    import shutil, psutil, glob, csv
    from zipfile import ZipFile
    import win32com.client as win32
    import xlwings as xw
    import xlsxwriter
    import xlrd
    import openpyxl
    import openpyxl.reader.excel as openpyxl_Excel
    import openpyxl.styles as styl
    #from openpyxl.styles import NamedStyle, Font, PatternFill, colors, Border, Side # , Alignment, Color
    import fct_dataframe as dframe
except Exception as err:
    str_lib = str(err).replace("No module named ", "").replace("'", "")
    print(" ATTENTION,  Missing library: '{0}' \n * Please Open Anaconda prompt and type: 'pip install {0}'".format(str_lib))


def fStr_Message(str_in):
    print(str_in)
    return '\n' + str_in


#---------------------------------------------------------------
#------------- Decorator -------------
#---------------------------------------------------------------
def dec_singletonsClass(input_classe):
    '''
    Singeltons decorators: prendre toujours que la premiere instance 
    (exemple : instance de connexion a database, on ne veut pas plusieurs instances
    , mais tjrs la premiere si elle existe)
    '''    
    d_instances = {}
    def wrap_getInstances(*l_paramInput, **d_paramInput):
        if input_classe not in d_instances:
            # Add instances as value in the dictionary where the key is the class
            d_instances[input_classe] = input_classe(*l_paramInput, **d_paramInput)
        # If an instance already exist for ones class, just use this instance
        return d_instances[input_classe]
    return wrap_getInstances


def dec_getTimePerf(int_secondesLimitDisplay = 1):
    '''
    Time Performance Decorators on a function
    You can calculate and compare Performance on any function just by decorating it
    '''    
    def dec_decoratorinside(input_fct):
        def wrap_modifiedFunction(*l_paramInput, **d_paramInput):
            # Before Function Execution...
            time_Debut = time.time()
            # Function execution 
            #   If you want to make stuff after execution of the function, you need to call function before returning it    
            launchFunction = input_fct(*l_paramInput, **d_paramInput)
            # After Function Execution...
            time_Fin = time.time()
            time_duree = time_Fin - time_Debut
            sec_duree = int(time_duree)
            milli_duree = int((time_duree - sec_duree) * 1000)
            if sec_duree >= int_secondesLimitDisplay:
                print(' * Execution time: {} = {} sec, {} milliSec'.format(input_fct, sec_duree, milli_duree))
            # Return the Function at the end
            return launchFunction
        return wrap_modifiedFunction
    return dec_decoratorinside


#---------------------------------------------------------------
# ------------- CLASS Excel Management (XLS, XLSX) -------------
#---------------------------------------------------------------
@dec_singletonsClass
class c_xlApp_xlwings():
    '''# DOC: https://docs.xlwings.org/en/stable/api.html'''
    def __init__(self):
        self.__wkIsOpen = False
        self.d_wkOpen = {}
        self.fBl_ExcelIsOpen()
        self.__int_xlsAppPid = -1
        self.xl_lastSheet = None
    
    #=====================================================        
    @property
    def visible(self):
        return self.__visible
    @visible.setter
    def visible(self, bl_visible):
        self.__visible = bl_visible
        self.xl_App.visible = self.__visible
    @property
    def screen_updating(self):
        return self.__screen_updating
    @screen_updating.setter
    def screen_updating(self, bl_screen_updating):
        self.__screen_updating = bl_screen_updating
        self.xl_App.screen_updating = self.__screen_updating
    @property
    def display_alerts(self):
        return self.__display_alerts
    @display_alerts.setter
    def display_alerts(self, bl_display_alerts):
        self.__display_alerts = bl_display_alerts
        self.xl_App.display_alerts = self.__display_alerts
    @property
    def wb_path(self):
        return self.__wb_path
    @wb_path.setter
    def wb_path(self, str_path):
        self.__wb_path = str_path
    #=====================================================
    
    def fBl_ExcelIsOpen(self):
        try:
            xl_Apps = xw.apps
            if xl_Apps: self.__blXlWasOpen = True
            else:       self.__blXlWasOpen = False
        except Exception as err: 
            self.__blXlWasOpen = False
            print('  WARNING in fBl_ExcelIsOpen: {}'.format(str(err)))
        return self.__blXlWasOpen
        
    def FindXlApp(self, bl_visible = True, bl_screen_updating = True, bl_display_alerts = True):
        '''Always get a new instance of Excel that we are going to kill'''        
        # Additonal check to be sure
        xl_Apps = xw.apps
        if not xl_Apps:
            self.__int_xlsAppPid = -1
        
        # Get the App
        if self.__int_xlsAppPid >= 0:
            try:
                xl_App = xl_Apps[int(self.__int_xlsAppPid)]
            except:
                xl_App = None
                try:
                    for App in xl_Apps:
                        pid = App.pid
                        if pid == self.__int_xlsAppPid:
                            xl_App = App
                    if xl_App is None: raise
                except:
                    xl_App = xl_Apps.active
                    if self.__int_xlsAppPid == xl_App.pid:
                        print(' INFORMATION: pid method did not work')
                    else:
                        print(" WARNING: We saved the PID as #{}# but active pid is #{}#".format(self.__int_xlsAppPid, xl_App.pid))
                        print(' - List of Excel App open: {}'.format(xl_Apps))
                        self.__int_xlsAppPid = xl_App.pid
        else:
            try:    
                xl_App = xw.apps.add()   # or xl_App = xw.App()
                self.__int_xlsAppPid = int(xl_App.pid)
            except Exception as err:
                print('  ERROR in FindXlApp || {}'.format(str(err)))
                raise
        # Options
        self.__visible =        bl_visible
        self.__screen_updating= bl_screen_updating
        self.__display_alerts = bl_display_alerts
        xl_App.visible =         self.__visible
        xl_App.screen_updating = self.__screen_updating
        xl_App.display_alerts =  self.__display_alerts
        self.xl_App = xl_App
        return self.xl_App
    
    def OpenWorkbook(self, str_path = '', str_password = None):
        if str_path != '':          self.wb_path = str_path            
        # Open Workbook
        #xl_book = xw.Book(str_path...)
        try:
            xl_book = self.xl_App.books.open(self.wb_path)
        except Exception as err:    print(' ERROR in OpenWorkbook - Open || {}'.format(str(err)))
        # Fill PARAM
        self.xl_lastBook = xl_book
        self.__wkIsOpen = True
        # Dico - {path : obj_workbook}
        self.d_wkOpen[self.wb_path] = xl_book
        # Just for the record
        #        xl_books = self.xl_App.books # All Workbooks open (xl_books = xw.books  # on active app)
        #        xl_book = self.xl_App.books['Book1'] # Select one workbook by its name (xl_book = xw.Book('Book1'))
        #        xl_book = xl_books.add() # Create a book    (xl_book = xw.Book())
        return self.xl_lastBook
    
    # SHEET
    def DefineWorksheet(self, str_sheetName = '', int_sheetNumber = -1, str_sheetNameToADD = ''):
        xl_lastBook = self.xl_lastBook
        self.xl_lastSheet_name = str_sheetName
        
        # Sheet with a Name
        if str_sheetName != '':	
            try:	xl_sheet = xl_lastBook.sheets[str_sheetName]
            except:							# After an error, all should have been defined in the call, so RETURN to get out
                self.DefineWorksheet('', int_sheetNumber, str_sheetNameToADD = str_sheetName)
                return self.xl_lastSheet
        else:
            # Sheet with a Number
            if int_sheetNumber > 0:
                try:
                    xl_sheet = xl_lastBook.sheets[int_sheetNumber - 1]
                    #int_sheetNumber = xl_sheet.index
                except:
                    self.DefineWorksheet('', -1, str_sheetNameToADD)
                    return self.xl_lastSheet
                # Rename the Sheet if possible (if its a recall with str_sheetName defined in str_sheetNameToADD)
                self.xl_lastSheet = xl_sheet  #(ONly for the rename)
                self.RenameSheet(str_sheetNameToADD)
            else:
                # ADD worksheet
                self.AddWorksheet(str_sheetNameToADD)
                return self.xl_lastSheet 		# All defined in Add worksheet, we can get out
        # End
        self.xl_lastSheet = xl_sheet
        self.SelectWorksheet()
        return self.xl_lastSheet
    
    def RenameSheet(self, str_sheetName = ''):
        try:
            if str_sheetName != '':
                self.xl_lastSheet.name = str_sheetName
                self.xl_lastSheet_name = str_sheetName
        except Exception as err:      print('  ERROR in Rename Sheet, sh Name: #{}# || {}'.format(str_sheetName, str(err)))
        
    def AddWorksheet(self, str_sheetName = None, str_beforeSheet = None, str_afterSheet = None):
        xl_lastBook = self.xl_lastBook
        xl_sheets = xl_lastBook.sheets
        try:
            xl_sheet = xl_sheets.add(name = str_sheetName, before = str_beforeSheet, after = str_afterSheet)
        except Exception as err:
            print(' ERROR in AddWorksheet. Sheet Name: #{}# || {}'.format(str_sheetName, str(err)))
        self.xl_lastSheet = xl_sheet
        # Define Sheet Name
        self.xl_lastSheet_name = str_sheetName
        self.SelectWorksheet()
        return self.xl_lastSheet
    
    def SelectWorksheet(self):
        xl_lastSheet = self.xl_lastSheet
        try:
            xl_lastSheet.select()
            #xl_lastSheet.activate
            return True
        except Exception as err:
            print(' ERROR in SelectWorksheet. Sheet Name = #{}# || {}'.format(self.xl_lastSheet_name, str(err)))
        return False
        
    def InsertDf_inRange(self, df, t_cell = (1,1), bl_autofitCol = False):
        xl_sheet = self.xl_lastSheet
        #https://docs.xlwings.org/en/stable/converters.html
        if t_cell == (1,1):
            xl_sheet.range('A1').options(index = False, header = False).value = df
        else:
            xl_sheet.range(t_cell).options(index = False, header = False).value = df
        # Just for the record
        #        xl_sheet.Range('A1:C3')
        #        xl_sheet.Range((1,1), (3,3))
        #        xl_sheet.Range('NamedRange')
        #        xl_sheet.Range(xw.Range('A1'), xw.Range('B2'))
        #        xl_sheet.clear()
        #        xl_sheet.clear_contents()
        #        xl_sheet.delete()

        # Autofits the width of either columns, rows or both on a whole Sheet.
        if bl_autofitCol:
            xl_sheet.autofit('columns')
        return True
    
    def save_Book(self, str_newPath = ''):
        xl_book = self.xl_lastBook
        if self.__wkIsOpen:
            # Display Alert
            self.xl_App.DisplayAlerts = False
            try:
                if str_newPath == '':   xl_book.save()
                else:                   xl_book.save(str_newPath)
            except Exception as err:
                print('  Error in SaveAs (Files): {}'.format(str(err)))
                raise
            finally:
                self.xl_App.DisplayAlerts = self.display_alerts
        else:   print('  ERROR in Save_Book | a WB need to be open before to bes Saved')
	
    def close_Book(self, bl_saveBeforeClose = None):
        if not bl_saveBeforeClose is None:
            self.__saveBeforeClose = bl_saveBeforeClose
        if self.__wkIsOpen:
            xl_book = self.xl_lastBook
            # Save
            if self.__saveBeforeClose:
                self.save_Book()
            # Close without saving
            xl_book.close()     
            # check if still Wk Open 
            self.CheckAnyWkIsOpen()
        return True
            
    def Quit_xlApp(self, bl_force = False, bl_killExcelProcess = False):
        # Input
        self.__killExcelProcess = bl_killExcelProcess
        self.xl_App.visible =         True
        self.xl_App.screen_updating = True
        self.xl_App.display_alerts =  True
        # Force or not Force
        if bl_force:
            if self.__wkIsOpen:
                try:    self.close_Book()
                except: print(' ERROR on Quit_xlApp: close_Book did not work. We are in Force Close mode: We kill Excel anyway !')
            self.KillExcel()
        else:
            # if a previous Excel session App is open, not killing the process but only current App
            if self.__blXlWasOpen:
                if self.__killExcelProcess:
                    print('''  (*) Warning on Quit_xlApp: A previous workbook is still Open. 
                          Not Killing process but just Excel App
                          If you want to kill Excel Process, SET bl_force = True  AND bl_killExcelProcess = True''')
                    self.__killExcelProcess = False
            if self.__wkIsOpen:
                self.CheckAnyWkIsOpen(bl_message_lOpenBook = True)
                if self.__wkIsOpen:
                    print('  (*) Warning on Quit_xlApp: Not closing EXCEL, at least a workbook is still Open')
                else:
                    self.KillExcel()
            else:   self.KillExcel()
        return True

    def KillExcel(self):
        # Additonal Check if some Wk are open
        if self.__wkIsOpen:
            print('''  INFORMATION on misusage: you try to KillExcel while you still have Excel Workbook Open. 
                  It will be saved but code should not be used this way''')
            self.CheckAnyWkIsOpen(bl_message_lOpenBook = True)
            self.close_Book(bl_saveBeforeClose = True)
        # KillExcel
        try:
            if self.__killExcelProcess:
                for proc in psutil.process_iter():
                    if any(procstr in proc.name() for procstr in ['Excel', 'EXCEL', 'excel']):
                        proc.kill()
            else:
                self.xl_App.quit() 
                self.xl_App.kill()
            #----- restart Init for the Next Instance -----
            self.__wkIsOpen = False
            self.d_wkOpen = {}
            self.__blXlWasOpen = False
            self.__int_xlsAppPid = -1
        except Exception as err:
            print('  Error in KillExcel: {}'.format(str(err)))
            raise
            
    def CheckAnyWkIsOpen(self, bl_message_lOpenBook = False):
        try:
            xl_App = self.xl_App
            l_wkOpen = [xl_book.name for xl_book in xl_App.Workbooks]
            if bl_message_lOpenBook:    print(' List of open Workbook: {}'.format(l_wkOpen))
            d_wkOpen_Copy = self.d_wkOpen.copy()
            for path, xl_book in d_wkOpen_Copy.items():
                try:
                    xl_book_name = xl_book.name
                    if not xl_book_name in l_wkOpen:
                        del self.d_wkOpen[path]
                except: del self.d_wkOpen[path]
            # Conclude if any wk is open
            if self.d_wkOpen:   self.__wkIsOpen = True
            else:               self.__wkIsOpen = False
        except: # Exception as err:
            #print('  INFORMATION: CheckAnyWkIsOpen (Files): {}'.format(str(err)))
            self.__wkIsOpen = False
            self.__killExcelProcess= True
        return self.__wkIsOpen
    
    #----------------------------------------- 
    # JUST for the record    
    #-----------------------------------------
    def ExecuteMacro(self, str_macroName, o_arg = None):
        vb_function = self.xl_App.macro(str_macroName)
        o_result = vb_function(o_arg)
        return o_result

    def CodeToTry():
        pass
        #import xlwings as xw
        #import pandas as pd
        #import numpy as np
        #str_path = r'C:\Users\Laurent.Tu\Documents\Taff\Nikko AM.xlsx'        
        #xl_Apps = xw.apps
        #if xl_Apps:
        #    xl_App = xl_Apps.active
        #else:
        #    xl_App = xl_Apps.add()        
        #xl_book = xl_App.books.open(str_path)
        #xl_sheets = xl_book.sheets
        #xl_sheet = xl_sheets[0]
        #df = pd.DataFrame(np.random.rand(10, 4), columns=['a', 'b', 'c', 'd'])
        #xl_sheet.range('A1').options(index = False, header = False).value = df
        #xl_book.save()
        #xl_book.close()
        #xl_App.quit()
        #xl_App.kill()



@dec_singletonsClass
class c_win32_xlApp():
    def __init__(self):
        #print('   *** Init class Excel Manager...')
        self.__wkIsOpen = False
        self.d_wkOpen = {}
        self.fBl_ExcelIsOpen()
    
    #=====================================================
    @property
    def visible(self):
        return self.__visible
    @visible.setter
    def visible(self, bl_visible):
        self.__visible = bl_visible
    @property
    def wb_path(self):
        return self.__wb_path
    @wb_path.setter
    def wb_path(self, str_path):
        self.__wb_path = str_path
    #=====================================================
    
    def fBl_ExcelIsOpen(self):
        try:
            self.xlApp = win32.GetActiveObject("Excel.Application")
            self.__blXlWasOpen = True
        except:
            self.__blXlWasOpen = False
        
    def FindXlApp(self, bl_visible = True, bl_gencache_EnsureDispatch = False):
        '''Get running Excel instance if possible, else return new instance.'''
        self.__visible = bl_visible
        self.__gencache_EnsureDispatch = bl_gencache_EnsureDispatch
        try:
             #print("Running Excel instance found, returning object")
            xlApp = self.xlApp
        except:
            #print("No running Excel instances, returning new instance")
            try:            
                if self.__gencache_EnsureDispatch:
                    xlApp = win32.gencache.EnsureDispatch('Excel.Application')
                else:
                    xlApp = win32.Dispatch('Excel.Application')
                    #xlApp = win32.DispatchEx('Excel.Application')
                    #xlApp = win32.dynamic.Dispatch('Excel.Application')
            except AttributeError as err_att:
                if "no attribute 'CLSIDToClassMap'" in str(err_att):
                    print('  WARNING in FindXlApp: no attribute CLSIDToClassMap || {}'.format(str(err_att)))
                    self.del_Gen_py_folder('FindXlApp')
                    return self.xlApp
                else:
                    print('  ERROR in FindXlApp || {}'.format(str(err_att)))
                    raise
        xlApp.Visible = self.__visible
        self.xlApp = xlApp
        return self.xlApp
		
    def WaitFile(self, int_sec = 1, str_msg = ' (*-*) Wait for file to load (in c_win32_xApp)...', otherARG = ''):
        if otherARG != '': print('otherARG', otherARG, type(otherARG))
        print(str_msg)
        time.sleep(int_sec)
        
    def del_Gen_py_folder(self, str_function):
        #=====================================================
        # Documentation on the subject:
        # https://gist.github.com/rdapaz/63590adb94a46039ca4a10994dff9dbe
        # https://stackoverflow.com/questions/47608506/issue-in-using-win32com-to-access-excel-file/47612742
        #=====================================================
        str_DirPath = os.path.join(os.environ['USERPROFILE'], r'AppData\Local\Temp\gen_py')
        print('   (***) delete folder : {}'.format(str_DirPath))
        if fBl_FolderExist(str_DirPath):
            # Delete the folder
            shutil.rmtree(str_DirPath, ignore_errors = True)
        # Re- Launch Process
        if str_function == 'FindXlApp':
            # Define again the App
            xlApp = win32.Dispatch('Excel.Application')
            self.xlApp = xlApp
            return self.xlApp
        
    def OpenWorkbook(self, str_path = '', str_password = ''):
        if str_path != '':          self.wb_path = str_path            
        # OPEN
        if str_password != '':
            xlWb = self.xlApp.Workbooks.Open(self.wb_path, False, True, None, Password = str_password)
        else:
            xlWb = self.xlApp.Workbooks.Open(self.wb_path)
        self.xl_lastWk = xlWb
        # Dico - {path : obj_workbook}
        self.d_wkOpen[self.wb_path] = xlWb
        self.__wkIsOpen = True
        return self.xl_lastWk
    
    
    def SelectWorksheet(self):
        xlWs = self.xl_lastWsh
        # Authorize 10 try to add worksheet
        for i_try in range(1, 6):
            try:
                xlWs.Select
                return True
            except:		self.WaitFile(1, ' (**) Error on SelectWorksheet (in c_win32_xApp), try number {}'.format(str(i_try)))
        return False
				
    def AddWorksheet(self, str_sheetName = ''):
        xlWb = self.xl_lastWk
        # Authorize 10 try to add worksheet
        for i_try in range(1, 6):
            try:
                if str_sheetName == '':		xlWs = xlWb.add_worksheet()
                else:						xlWs = xlWb.add_worksheet(str_sheetName)
                break
            except:		self.WaitFile(1, ' (**) Error on AddWorksheet (in c_win32_xApp), try number {}'.format(str(i_try)))
        self.xl_lastWsh = xlWs
        #--------------------------------------------------
        # Shitty stuff because sheet name is not recognised
        try:        self.str_lastSheetName = self.xl_lastWsh.Name
        except Exception as Err:
            print('  ERROR in  AddWorksheet(c_win32_xApp): {}'.format(str(Err)))
            print('  - str_sheetName: ', str_sheetName)
            self.str_lastSheetName = str_sheetName
        #--------------------------------------------------
        self.SelectWorksheet()
        return self.xl_lastWsh
	
    def RenameSheet(self, str_sheetName = ''):
        try:
            if str_sheetName != '':
                try:	self.xl_lastWsh.Name = str_sheetName  	        #xlWs.title
                except:	self.WaitFile(1, ' (**) Warning on DefineWorksheet: Could not rename Sheet into : {}'.format(str_sheetName))
        except Exception as err:      print('  ERROR in RenameSheet || {}'.format(str(err)))
        
    def DefineWorksheet(self, str_sheetName = '', int_sheetNumber = -1, str_sheetNameToADD = ''):
        xlWb = self.xl_lastWk
		# Sheets(str_sheetName)
		# Worksheets(str_sheetName)
		# ActiveSheet
		
        if str_sheetName != '':				# Name is defined
            try:	xlWs = xlWb.Sheets(str_sheetName)
            except:							# After an error, all should have been defined in the call, so RETURN to get out
                self.WaitFile(1, ' (**) Warning on DefineWorksheet: Could not find Sheet Name : {}'.format(str_sheetName))
                self.DefineWorksheet('', int_sheetNumber, str_sheetNameToADD = str_sheetName)
                return self.xl_lastWsh
            self.xl_lastWsh = xlWs
        else:
            if int_sheetNumber > 0:         # Number is defined
                try:	
                    xlWs = xlWb.Sheets(int_sheetNumber)
                except:						# After an error, all should have been defined in the call to add worksheet, so RETURN to get out
                    self.WaitFile(1, ' (**) Warning on DefineWorksheet: Could not find Sheet Number : {}'.format(str(int_sheetNumber)))
                    self.DefineWorksheet('', -1, str_sheetNameToADD)
                    return self.xl_lastWsh
                self.xl_lastWsh = xlWs
                # Rename the Sheet if possible (if its a recall with str_sheetName defined in str_sheetNameToADD)
                self.RenameSheet(str_sheetNameToADD)
            else:
                # ADD worksheet
                self.AddWorksheet(str_sheetNameToADD)
                return self.xl_lastWsh 		# All defined in Add worksheet, we can get out
        # End
        #--------------------------------------------------
        # Shitty stuff because sheet Name is not recognised
        try:    self.str_lastSheetName = self.xl_lastWsh.Name
        except Exception as Err:
            print('  ERROR in  DefineWorksheet(c_win32_xApp): {}'.format(str(Err)))
            print('  - ', str_sheetName, int_sheetNumber, str_sheetNameToADD)
            self.str_lastSheetName = str_sheetName
        #--------------------------------------------------
        self.SelectWorksheet()
        return self.xl_lastWsh
    
    def SaveAs(self, str_newPath, int_fileFormat = -1):
        if self.__wkIsOpen:
            # Define FileFormat
            str_lower = str_newPath.lower()
            if int_fileFormat == -1:
                if '.xlsx' in str_lower:        self.__fileFormat = 51
                elif '.xlsb' in str_lower:      self.__fileFormat = 50
                elif '.xlsm' in str_lower:      self.__fileFormat = 52
                elif '.xls' == str_lower[-4:]:  self.__fileFormat = 56
                else:                           self.__fileFormat = -1
            else:                               self.__fileFormat = int_fileFormat
            # Save As
            self.__displayAlert = self.xlApp.DisplayAlerts
            self.xlApp.DisplayAlerts = False
            try:
                if self.__fileFormat == -1:		self.xl_lastWk.SaveAs(str_newPath)
                else:							self.xl_lastWk.SaveAs(str_newPath, FileFormat = self.__fileFormat)
            except Exception as err:
                print('  Error in SaveAs (Files): {}'.format(str(err)))
                raise
            finally:
                self.xlApp.DisplayAlerts = self.__displayAlert
        else:   print('  ERROR in SaveAs (c_win32_xApp) | a WB need to be open before to bes Saved AS')
	
    def CloseWorkbook(self, bl_saveBeforeClose = True):
        self.__saveBeforeClose = bl_saveBeforeClose
        if self.__wkIsOpen:
            self.xl_lastWk.Close(SaveChanges = self.__saveBeforeClose)
		
    def CheckAnyWkIsOpen(self):
        try:
            if self.__gencache_EnsureDispatch:		xlApp2 = win32.gencache.EnsureDispatch('Excel.Application')
            else:									xlApp2 = self.xlApp
            l_wkOpen = [wk.Name for wk in xlApp2.Workbooks]
            #            d_wkOpen = {path: wk for path, wk in self.d_wkOpen.items() if wk.Name in l_wkOpen}
            #            self.d_wkOpen = d_wkOpen
            
            d_wkOpenCopy = self.d_wkOpen.copy()
            for path, wk in d_wkOpenCopy.items():
                try:
                    wk_Name = wk.Name
                    if not wk_Name in l_wkOpen:
                        del self.d_wkOpen[path]
                except: del self.d_wkOpen[path]
            # Conclude if any wk is open
            if self.d_wkOpen:   self.__wkIsOpen = True
            else:               self.__wkIsOpen = False
        except: # Exception as err:
            #print('  INFORMATION: CheckAnyWkIsOpen (Files): {}'.format(str(err)))
            self.__wkIsOpen = False
            self.__killExcelProcess= True
        return self.__wkIsOpen
    
    def Kill_Excel(self):
        try:
            if self.__killExcelProcess:
                for proc in psutil.process_iter():
                    if any(procstr in proc.name() for procstr in ['Excel', 'EXCEL', 'excel']):
                        proc.kill()
            #-----------------------------------------------------------------------
            elif False:     # With win32com
                o_WbemScripting = win32.Dispatch("WbemScripting.SWbemLocator")
                o_cimv2 = o_WbemScripting.ConnectServer(".", "root\cimv2")
                o_excelProcess = o_cimv2.ExecQuery("Select Caption from Win32_Process where Caption LIKE 'EXCEL%'")
                for excel in o_excelProcess:
                    try:    excel.terminate()
                    except: pass
            elif False:     # Other things picked up on Internet
                if hasattr(self, 'xlBook'):
                    print(' WARNING in Kill_Excel (Files): remaining xlBook.....')
                    del self.xl_lastWk
                import gc
                gc.collect()
            #-----------------------------------------------------------------------
            else:
                self.xlApp.Application.Quit()
                del (self.xlApp)
            #----- restart Init for the Next Instance -----
            self.__wkIsOpen = False
            self.d_wkOpen = {}
            self.__blXlWasOpen = False
        except Exception as err:
            print('  Error in Kill_Excel (Files): {}'.format(str(err)))
            raise
        
    def QuitXlApp(self, bl_force = False, bl_killExcelProcess = False):
        self.__killExcelProcess = bl_killExcelProcess
        if bl_force:
            if self.__wkIsOpen:
                try:    self.CloseWorkbook()
                except: pass
            self.Kill_Excel()
        else:
            if self.__blXlWasOpen:
                print('  (*) Warning QuitXlApp(c_win32_xApp): Not closing EXCEL, a previous workbook might be still Open')
            else:
                self.CheckAnyWkIsOpen()
                if self.__wkIsOpen:
                    print('  (*) Warning QuitXlApp(c_win32_xApp): Not closing EXCEL, a workbook is still Open')
                else:       
                    self.Kill_Excel()



# KILL EXCEL
def Act_KillExcel():
    for proc in psutil.process_iter():
        if any(procstr in proc.name() for procstr in ['Excel', 'EXCEL', 'excel']):
            proc.kill()
    
    
#------------------------------------------------------------------------------
# List Files in folder
#------------------------------------------------------------------------------
def fStr_GetFolderFromPath(str_path):
    str_folder = str('\\'.join(str_path.split('\\')[:-1]))
    return str_folder
def fStr_GetFileFromPath(str_path):
    str_fileName = str(str_path.split('\\')[-1])
    return str_fileName
    
def fStr_BuildPath(str_folder, str_FileName):
    if str_FileName == '':      str_path = str_folder
    elif str_folder == '':      str_path = str_FileName
    else:                       str_path = os.path.join(str_folder, str_FileName)
    return str_path

def fStr_BuildFolder_wRoot(str_folderPart, str_folderRoot):
    if str_folderPart[:2] == '\\\\':    return str_folderPart
    elif str_folderPart[:2] == 'C:':    return str_folderPart
    elif str_folderPart[:2] == 'E:':    return str_folderPart
    elif 'Manual_py' in str_folderPart:
        return str_folderRoot.replace('Auto_py\\', '') + str_folderPart
    else:                               
        return str_folderRoot + str_folderPart
    
def fBl_FileExist(str_path):
    if os.path.isfile(str_path):    return True
    else:                           return False

def fBl_FolderExist(str_path):
    if os.path.exists(str_path):    return True
    else:                           return False
    
def fList_FileInDir(str_path):
    try:        l_fic = os.listdir(str_path)
    except:
        print(' ERROR in fList_FileInDir')
        print(' - ', str_path)
        raise
    return l_fic

def fList_FileInDir_Txt(str_path):
    try:
        l_fic = os.listdir(str_path)
        l_fic = [fic for fic in l_fic if '.txt' in fic]
    except:
        print(' ERROR in fList_FileInDir_Txt')
        print(' - ', str_path)
        raise
    return l_fic

def fList_FileInDir_Csv(str_path):
    try:
        l_fic = os.listdir(str_path)
        l_fic = [fic for fic in l_fic if '.csv' in fic]
    except:
        print(' ERROR in fList_FileInDir_Csv')
        print(' - ', str_path)
        raise
    return l_fic

def fList_FileInDir_Py(str_path):
    try:
        l_fic = os.listdir(str_path)
        l_fic = [fic for fic in l_fic if '.py' in fic]
    except:
        print(' ERROR in fList_FileInDir_Py')
        print(' - ', str_path)
        raise
    return l_fic

def UpdateTxtFile(str_path, str_old, str_new = ''):
    with open(str_path, 'r') as file :
        str_text = file.read()
    # Replace the target string
    str_text = str_text.replace(str_old, str_new)
    # Write the file out again
    with open(str_path, 'w') as file:
        file.write(str_text)


#------------------------------------------------------------------------------
# Open Files
#------------------------------------------------------------------------------
def fBk_OpenWk_xlrd(str_path):
    o_Book = xlrd.open_workbook(str_path)
    return o_Book


#------------------------------------------------------------------------------
# Transform Names
#------------------------------------------------------------------------------
def Act_Rename(str_folder, str_OriginalName, str_NewName, bl_message = True):    
    try:
        if str_NewName.upper() != str_OriginalName.upper():
            if bl_message:
                print(' RENAMING')
                print(' - str_OriginalName    : ', str_OriginalName)
                print(' - str_NewName: ', str_NewName)
            try:        os.rename(os.path.join(str_folder, str_OriginalName), os.path.join(str_folder, str_NewName))
            except:     shutil.move(os.path.join(str_folder, str_OriginalName), os.path.join(str_folder, str_NewName))
    except: 
        print(' ERROR in Act_Rename: Rename stuff')
        print(' - str_folder: ', str_folder)
        print(' - str_OriginalName: ', str_OriginalName)
        print(' - str_NewName: ', str_NewName)
        raise
    return True


def fStr_TransformFilName_fromXXX_forGlobFunction(str_fileName_withX, bl_exactNumberX):
    # Check if its a normal Name without {X}:
    if '{X' not in str_fileName_withX and 'X}' not in str_fileName_withX:
        return str_fileName_withX
    
    # Count the Number of Series of {XX} 
    int_nbXX = str_fileName_withX.count('{X')
    int_nbXX2 = str_fileName_withX.count('X}')
    if int_nbXX != int_nbXX2: 
        print('   ERROR, check the sting str_fileName_withX in fStr_TransformFilName_fromXXX: ', str_fileName_withX)
        return str_fileName_withX
    
    # Count the number of X in each Series of {XX}
    str_fileName = str_fileName_withX
    nb = 1
    while nb in range(1, int_nbXX + 1):
        nb += 1
        for i in range(1,100):
            str_XXX = '{' + i * 'X' + '}'
            if str_fileName.count(str_XXX) > 0:
                nb = nb + str_fileName.count(str_XXX) - 1     # just in case there is several time the same XXX, we dont want to pass again on this loop
                break                
        #==================================================
        # Exact Number ???????????????????????????????
        if bl_exactNumberX:
            int_lenXX = len(str_XXX) - 2
            str_fileName = str_fileName.replace(str_XXX, int_lenXX * '?')
        # Flex Number ?
        else:
            str_fileName = str_fileName.replace(str_XXX, '*')
        #==================================================
    return str_fileName



#Return a list of File in a folder
def fL_GetFileListInFolder(str_folder, str_fileName_withX, bl_searchOnlyIfPossible = False, bl_exactNumberX = True):    
    if str_folder[-1] != '\\': str_folder += '\\'
    
    try:
        # Transform fiel name to be understood from 'glob.glob'
        str_fileName = fStr_TransformFilName_fromXXX_forGlobFunction(str_fileName_withX, bl_exactNumberX)
        # if no change, just return default
        if str_fileName == str_fileName_withX:
            return [str_folder + str_fileName_withX]
        
        # Using Glob: Code part from Thanos
        for file in [glob.glob(str_folder + str_fileName)]:
            if len(file) > 0:
                L_files = glob.glob(str_folder + str_fileName)
                #++++++++++++++++++++++++++++++++++++++++++++++++++++
                return L_files
                #++++++++++++++++++++++++++++++++++++++++++++++++++++
            # Do not raise an issue if its a search, just return File with XX
            elif bl_searchOnlyIfPossible:                   
                #print(' Return the file with the X (Search only): ', str_fileName_withX)
                return [str_folder + str_fileName_withX]
            else:
                if bl_exactNumberX:
                    print(' EMPTY fL_GetFileListInFolder... We will now search with more flex')
                    l_filesFlex = fL_GetFileListInFolder(str_folder, str_fileName_withX, False, False)
                    return l_filesFlex
                else:                
                    print(' EMPTY: file not found in fL_GetFileListInFolder')
                    raise
    except:
        print('   ERROR in fL_GetFileListInFolder')
        print('   - str_folder: ', str_folder)
        print('   - str_fileName_withX: ', str_fileName_withX)
        raise
    return 0 #never used
        

def fStr_GetMostRecentFile_InFolder(str_folder, str_fileName_withX, bl_searchOnlyIfPossible = False, bl_exactNumberX = True):
    if str_folder[-1] != '\\': str_folder += '\\'
    # Get the lisyt of matching files
    try:
        L_FileInFolder = fL_GetFileListInFolder(str_folder, str_fileName_withX, bl_searchOnlyIfPossible, bl_exactNumberX)
    except: raise
    # Return the most recent file
    try:
        #str_latest_file = max(L_FileInFolder, key = os.path.getmtime)       # Time of Update
        str_latest_file = max(L_FileInFolder, key = os.path.getctime)       # Time of Creation
        str_fileName = str_latest_file.replace(str_folder, '')
    except:
        if bl_searchOnlyIfPossible: return str_fileName_withX
        print('  ERROR fStr_GetMostRecentFile_InFolder, Sort by Date')
        print('  - str_folder: ',str_folder)
        print('  - str_fileName_withX: ',str_fileName_withX)
        raise
    return str_fileName

## TEST
#print(fStr_GetMostRecentFile_InFolder(r'C:\Users\laurent.tupin\Documents\5_Python', 'test_2152_5412.txt'))
#print(fStr_GetMostRecentFile_InFolder(r'C:\Users\laurent.tupin\Documents\5_Python', 'test_{XXXX}_5412.txt'))
#print(fStr_GetMostRecentFile_InFolder(r'C:\Users\laurent.tupin\Documents\5_Python', 'test_{XXXX}_{XX}12.txt'))


def fL_GetFileList_withinModel(L_FileName, str_fileName_withX):
    try:
        # Check if its a normal Name without {X}:
        if '{X' not in str_fileName_withX and 'X}' not in str_fileName_withX:
            L_FileName = [fil for fil in L_FileName if str_fileName_withX.lower() in fil.lower()]
            return L_FileName
			
        # Count the Number of Series of {XX} 
        int_nbXX = str_fileName_withX.count('{X')
        int_nbXX2 = str_fileName_withX.count('X}')
        if int_nbXX != int_nbXX2: 
            print('   ERROR, check the string str_fileName_withX in fL_GetFileList_withinModel: ', str_fileName_withX)
            return L_FileName
		
        # Count the number of X in each Series of {XX}
        str_fileName = str_fileName_withX
        for nb in range(1, int_nbXX + 1):
            for i in range(1,15):
                str_XXX = '{' + i * 'X' + '}'
                if str_fileName.count(str_XXX) == 1: break
            str_fileName = str_fileName.replace(str_XXX, '?')
			
        l_fileName_part = str_fileName.split('?')
        l_files = L_FileName
        for name_part in l_fileName_part:
            l_files = [fil for fil in l_files if name_part.lower() in fil.lower()]
    except:
        print('   ERROR in fL_GetFileList_withinModel')
        print('   - str_fileName_withX: ', str_fileName_withX)
        print('   - L_FileName: ', L_FileName)
        raise
    return l_files


#def fStr_FindPath_byReplacingX(str_folder, str_fileName_withX, l_fileInput = [], bl_searchOnlyIfPossible = False):
#    try:
#        # 1. List file in Folder
#        if l_fileInput: l_files = l_fileInput
#        else:           l_files = fList_FileInDir(str_folder)
#        
#        # 2. Understand the Name of the file
#        for i in range(1,15):
#            str_XXX = '{' + i * 'X' + '}'
#            if str_fileName_withX.count(str_XXX) == 1:
#                break
#            #elif str_fileName_withX.count(str_XXX) > 1:
#            #    print(' ERROR: fStr_FindPath_byReplacingX is not yet made to support Two {XXXX} string, you need to require an Upgrade of the function')
#        
#        #----------------------------------------------------------------------
#        # If it is a Normal File without XXX: Just return the input name
#        if str_fileName_withX.count(str_XXX) == 0: return str_fileName_withX
#        #----------------------------------------------------------------------
#        
#        # Treat the filename
#        str_prefix = str_fileName_withX.split(str_XXX)[0]
#        str_suffix = str_fileName_withX.split(str_XXX)[1]
#        # Loop on files in folder to get the name
#        l_files = [file for file in l_files if str_prefix in file and str_suffix in file]
#        # RETURN
#        if len(l_files) == 1:
#            return l_files[0]        
#        elif len(l_files) == 0:
#            #------------------------------------------------------------------
#            # If search only for the name of the already downloadedm dont retrun an error, but the original name
#            if bl_searchOnlyIfPossible: return str_fileName_withX
#            
#            print(' EMPTY: file not found in fStr_FindPath_byReplacingX')
#            print(' Maybe check the Date (file exist but not the right date, or the number of X exceed 15 ???? ')
#            raise
#        elif len(l_files) > 1:
#            for file in l_files:
#                fileName = file.replace(str_prefix, '').replace(str_suffix, '')
#                if len(fileName) == len(str_XXX) - 2:
#                    return file
#    except:
#        print('   ERROR in fStr_FindPath_byReplacingX')
#        print('   str_folder: ', str_folder)
#        print('   str_fileName_withX: ', str_fileName_withX)
#        raise




#------------------------------------------------------------------------------
# Files Date
#------------------------------------------------------------------------------ 
def fDte_GetModificationDate(str_pathFile):
    try:
        if fBl_FileExist(str_pathFile):
            dte_modif = os.path.getmtime(str_pathFile)
            dte_modif = dt.datetime.fromtimestamp(dte_modif)
            return dte_modif
        else:
            print('  fDte_GetModificationDate: File does not exist: ')
            print('  - str_pathFile: ', str_pathFile)
            return -1
    except:
        print(' ERROR in fDte_GetModificationDate')
        print(' - ', str_pathFile)
        raise
    return True


def fL_KeepFiles_wTimeLimit(l_pathFile, dte_after = 10, dte_before = 0):
    # Parameters in
    if type(dte_after) == int:
        dte_after = dt.datetime.now() - dt.timedelta(dte_after)
    if type(dte_before) == int:
        if dte_before != 0:
            dte_before = dt.datetime.now() - dt.timedelta(dte_before)
    # Keep file in list within the Limit Date
    try:
        l_pathReturn = [path for path in l_pathFile if fBl_FileExist(path)]
        l_pathReturn = [path for path in l_pathReturn if dt.datetime.fromtimestamp(os.path.getmtime(path)) > dte_after]
        if dte_before != 0:
            l_pathReturn = [path for path in l_pathReturn if dt.datetime.fromtimestamp(os.path.getmtime(path)) <= dte_before]
    except Exception as err:
        print(' ERROR in fL_KeepFiles_wTimeLimit: {}'.format(str(err)))
        raise
    return l_pathReturn



#-----------------------------------------------------------------
# CREATE
#-----------------------------------------------------------------
def fBl_createDir(str_folder):
    try:
        if not os.path.exists(str_folder):
            try: 
                os.makedirs(str_folder)
                return True
            except:
                print(' ERROR: fBl_createDir - Program could NOT create the Dir')
                print(' - ', str_folder)
                raise     
    except:
        print(' ERROR: fBl_createDir - Path cannot be tested on its existence ??')
        print(' - ', str_folder)
        raise 
    return False    # Folder already exist
    

def act_createFile(bl_folderRelative, str_folder, str_fileName = 'test.txt', str_text = ''):
    dir_current = os.getcwd()
    try:
        # Define folder
        if bl_folderRelative:
            str_folder = os.getcwd().replace(str_folder,'') + str_folder
        # Create folder
        fBl_createDir(str_folder)
        # Change Dir
        os.chdir(str_folder)
    except: 
        print(" ERROR: act_createFile - Create Dir")
        print(' - str_folder', str_folder)
        print(' - str_fileName', str_fileName)
        raise
    # Create File
    try:            f = open(str_fileName,"w+")
    except: 
        try:        os.chdir(dir_current)
        except:     print(" ERROR: act_createFile - os.chdir(dir_current) did not work -- f = open(str_fileName, ")
        print(" ERROR: act_createFile - Could not create the file")
        print(' - str_folder', str_folder)
        print(' - str_fileName', str_fileName)
        raise
    try: 
        f.write(str_text)
        f.close()
    except:
        try:        f.close()
        except:     print(" ERROR: act_createFile - f.close()")
        try:        os.chdir(dir_current)
        except:     print(" ERROR: act_createFile - os.chdir(dir_current) did not work -- f.write(str_text)")
        print(" ERROR: act_createFile - Could not write in the file")
        print(' - str_folder', str_folder)
        print(' - str_fileName', str_fileName)
        print(' - str_text', str_text)        
        raise
    try:            os.chdir(dir_current)
    except:
        print(" ERROR: act_createFile - os.chdir(dir_current) did not work")
        raise        
    return 0


def fStr_CreateTxtFile(str_folder, str_FileName, df_data, str_sep = '', bl_header = False, bl_index = False,
                       o_quoting = csv.QUOTE_MINIMAL, o_quotechar = '"'):
    try:
        if str_FileName == '':      str_path = str_folder
        else:                       str_path = os.path.join(str_folder, str_FileName)
        if str_sep == '':           str_sep = ','
        elif str_sep == '\\t':      str_sep = '\t'
        # TO CSV
        df_data.to_csv(str_path, sep = str_sep, 
                       header = bl_header, index = bl_index, 
                       quoting = o_quoting, #quotechar = o_quotechar
                       )
    except:
        print('  ERROR in fl.fStr_CreateTxtFile: Could not create the file: ')
        print('  - str_path :', str_path)
        print('  - other param :', str_sep, bl_header, bl_index, o_quoting)
        raise
    return str_path


#-----------------------------------------------------------------
# READ
#-----------------------------------------------------------------
def fStr_ReadFile_sql(str_filePath):
    # Open and read the file as a single buffer
    file = open(str_filePath, 'r')
    str_Content = file.read()
    file.close()
    return str_Content


def fStr_readFile(bl_folderRelative, str_folder, str_fileName = 'test.txt'):
    dir_current = os.getcwd()
    try:
        # Define folder
        if bl_folderRelative:
            str_folder = os.getcwd().replace(str_folder,'') + str_folder
        # Check folder exist
        if not os.path.exists(str_folder): 
            print(' ERROR: fStr_readFile - Folder does not exist')
            print(' - ', str_folder)
            print(' - ', str_fileName)
            print(' - Folder Relative = ', bl_folderRelative)
            raise
        # Change Dir
        os.chdir(str_folder)
    except: 
        print(" ERROR: fStr_readFile")
        print(' - ', str_folder, str_fileName)
        print(' - Folder Relative = ', bl_folderRelative)
        raise
    # Read File
    try:            f = open(str_fileName,"r")
    except: 
        try:        os.chdir(dir_current)
        except:     print(" ERROR: fStr_readFile - os.chdir(dir_current) did not work -- f = open(str_fileName")
        print(' ERROR: fStr_readFile')
        print(' - Could not Open the file')
        print(' - str_folder', str_folder)
        print(' - str_fileName', str_fileName)
        print(' - Folder Relative = ', bl_folderRelative)
        raise
    try:
        str_return = f.read()
        f.close()
    except:
        try:        f.close()
        except:     print(" ERROR: fStr_readFile - f.close()")
        try:        os.chdir(dir_current)
        except:     print(" ERROR: fStr_readFile - os.chdir(dir_current) did not work -- str_return = f.read()")
        print(" ERROR: fStr_readFile")
        print(" - Could not Read the file")
        print(' - str_folder', str_folder)
        print(' - str_fileName', str_fileName)
        print(' - Folder Relative = ', bl_folderRelative)
        raise
    try:            os.chdir(dir_current)
    except: 
        print(" ERROR: fStr_readFile - os.chdir(dir_current) did not work")
        raise     
    return str_return
    

#------------------------------------------------------------------------------
# DELETE
#------------------------------------------------------------------------------ 
def del_allTxtFileInFolder_ifOldEnought(str_folder, int_dayHisto = 60):
    try:
        dte_delete = dt.datetime.now() - dt.timedelta(int_dayHisto)
        l_fic = fList_FileInDir_Txt(str_folder)
        for fic in l_fic:
            str_path = os.path.join(str_folder, fic)
            if fDte_GetModificationDate(str_path) < dte_delete and fDte_GetModificationDate(str_path) != -1:
                os.remove(str_path)
    except:
        print(' ERROR in del_fic')
        print(' - ', str_folder, int_dayHisto)
        raise
    return 0


def del_fic(str_path, int_dayHisto):
    print('     DEPRECATED function: del_fic')
    a = del_allTxtFileInFolder_ifOldEnought(str_path, int_dayHisto)
    return a


def del_fichier(str_folder, str_fileName):
    try:
        str_path = os.path.join(str_folder, str_fileName)
        if str_fileName == '': str_path = str_folder
        os.remove(str_path)
    except:
        print(' ERROR in del_fichier')
        print(' - ', str_folder, str_fileName)
        raise
    return 0


def del_fichier_ifOldEnought(str_folder, str_fileName, int_dayHisto = 5):
    try:
        str_path = os.path.join(str_folder, str_fileName)
        if str_fileName == '': 
            str_path = str_folder
            str_folder = '\\'.join(str_folder.split('\\')[:-1])
        # if folder does not exist : sortir de la fonction sans delete rien (et en ayant creer le dossier)
        if fBl_createDir(str_folder):
            print(' Information: Folder was not existing (in del_fichier_ifOldEnought): ', str_folder)
            return 0
        dte_delete = dt.datetime.now() - dt.timedelta(int_dayHisto)
        dte_ModificationDate = fDte_GetModificationDate(str_path)
        if dte_ModificationDate == -1:
            # File not exisiting
            pass
        elif dte_ModificationDate < dte_delete:
            os.remove(str_path)
        else: pass
    except:
        print(' ERROR in del_fichier_ifOldEnought')
        print(' - Parameters:', str_folder, str_fileName, int_dayHisto)
        raise
    return 0
    


#-----------------------------------------------------------------
# CREATE XLS Files
#-----------------------------------------------------------------
def fStr_createExcel_1Sh(str_folder, str_FileName, df_Data, str_SheetName = '', bl_header = False):
    try:
        # Define Path
        str_path = fStr_BuildPath(str_folder, str_FileName)
        # Create the File
        if str_SheetName != '':     df_Data.to_excel(str_path, header = bl_header, index = False, sheet_name = str_SheetName)
        else:                       df_Data.to_excel(str_path, header = bl_header, index = False)
    except:
        print('  ERROR: fStr_createExcel_1Sh did not work ')
        print('  - str_path: ', str_path)
        print('  - str_SheetName: ', str_SheetName)
        return False
    return str_path


#options={'strings_to_numbers': True}
    
def fStr_createExcel_SevSh(str_folder, str_FileName, l_dfData, l_SheetName = [], bl_header = False, d_options = {}):
    try:
        # Define Path
        str_path = fStr_BuildPath(str_folder, str_FileName)
        # Create the File
        with pd.ExcelWriter(str_path, engine = 'xlsxwriter', options = d_options) as xl_writer:
            # Dataframe
            for i in range(len(l_dfData)):
                df_data = l_dfData[i]
                #Sheet Name
                try:        str_SheetName = l_SheetName[i]
                except:     str_SheetName = 'Sheet{}'.format(str(i + 1))
                # fill in
                df_data.to_excel(xl_writer, header = bl_header, index = False, sheet_name = str_SheetName)
            xl_writer.save()
    except:
        try:        xl_writer.save()
        except:     print('Could not Save the file')
        print('  ERROR: fl.fStr_createExcel_SevSh did not work ')
        print('  - ', str_folder, str_FileName)
        print('  - l_SheetName: ', l_SheetName)
        return False
    return str_path


def fStr_createExcel_SevSh_celByCel(str_folder, str_FileName, l_dfData, l_SheetName = [], bl_header = False):
    try:
        # Define Path
        if str_FileName == '':      str_path = str_folder
        else:                       str_path = os.path.join(str_folder, str_FileName)
        # Create the file (xlsxwriter cannot modify files)
        xlWb = xlsxwriter.Workbook(str_path)
        # Dataframe
        for i in range(len(l_dfData)):
            df_data = l_dfData[i]
            try:        str_SheetName = l_SheetName[i]
            except:     str_SheetName = ''
            #Sheet Name
            if str_SheetName != '':     xlWs = xlWb.add_worksheet(str_SheetName)             
            else:                       xlWs = xlWb.add_worksheet()   
            # fill in
            for i, row in enumerate(df_data.index):
                for j, col in enumerate(df_data.columns):
                    xlWs.write(i, j, str(df_data.iat[i, j]))
                    #xlWs.Cells(i+1, j+1).Value = str(df_data.iat[i, j])
        xlWb.close()
    except:
        try:        xlWb.close()
        except:     print('Could not close the file')
        print('  ERROR: fl.fStr_createExcel_SevSh_celByCel did not work ')
        print('  - ', str_folder, str_FileName)
        print('  - l_SheetName: ', l_SheetName)
        return False
    return str_path


# INSERT SHEET: 1 file out - 1 Dataframe - 1 Sheet 
def fStr_fillExcel_InsertNewSheet(str_folder, str_FileName, df_data, str_SheetName = '', bl_header = False):
    try:
        if str_FileName == '':  str_path = str_folder
        else:                   str_path = os.path.join(str_folder, str_FileName)
        # Define Book
        xl_book = openpyxl.load_workbook(filename = str_path)
        # OTHER Rows of code: 
        #   ws1 = xl_book.worksheets[ste_sheetName] 
        #   xl_book_2 = openpyxl.load_workbook(filename = str_path2)
        #   ws2 = xl_book_2.create_sheet(ws1.title)
        #   for row in ws1:
        #       for cell in row:
        #           ws2[cell.coordinate].value = cell.value
        #   xl_book_2.save(str_path2)
        # OTHER Rows of code: 
        #   xl_writer = pd.ExcelWriter(str_path, engine = 'openpyxl')
        
        with pd.ExcelWriter(str_path, engine = 'openpyxl') as xl_writer:
            xl_writer.book = xl_book
            xl_writer.sheets = dict((ws.title, ws) for ws in xl_book.worksheets)
            # Manege Sheet Name
            if str_SheetName == '':     str_SheetName = 'Sh1'
            if str_SheetName in list(xl_writer.sheets):
                print(" The sheet '{}' alrdeay exist".format(str_SheetName))
                while str_SheetName in list(xl_writer.sheets):
                    str_SheetName = str_SheetName[1:] + str_SheetName[0] + 'x'
                print(" We replace the sheet name: '{}'  but you need to improve the management of name".format(str_SheetName))
            # Add the Sheet
            df_data.to_excel(xl_writer, header = bl_header, index = False, sheet_name = str_SheetName)
            #SAVE
            xl_writer.save()
    except:
        print('  ERROR in fl.fStr_fillExcel_InsertNewSheet: Could not fill the file')
        print('  - str_path', str_path)
        print('  - str_SheetName', str_SheetName)
        return False
    return str_path


# 1 file out - n Dataframe - n Sheet
def fStr_fillXls_celByCel_plsSheets(str_folder,str_FileName,l_dfData,l_SheetName=[],l_nbRows=[],l_rowsWhere=[]):
    try:
        if str_FileName == '':      str_path = str_folder
        else:                       str_path = os.path.join(str_folder, str_FileName)        
        
        # Open the file (win32)
        inst_xlApp = c_win32_xlApp()
        inst_xlApp.FindXlApp(bl_visible = False)
        inst_xlApp.OpenWorkbook(str_path)
        #        xlApp = win32.Dispatch('Excel.Application')
        #        xlApp.Visible = False 
        #        xlWb = xlApp.Workbooks.Open(str_path)
		
        # Dataframe
        for i in range(len(l_dfData)):
            df_data = l_dfData[i]
            try:        str_SheetName = l_SheetName[i]
            except:     str_SheetName = ''
            
            #Sheet
            bl_SheetNameIsNew = (inst_xlApp.xl_lastWk.Sheets(i + 1).Name not in l_SheetName)
            if bl_SheetNameIsNew:   inst_xlApp.DefineWorksheet(str_SheetName, i + 1)
            else:                   inst_xlApp.DefineWorksheet(str_SheetName)
            
            # ------ Insert or delete ROWS ------
            int_nbRows = 0
            int_rowsWhere = 1
            if l_nbRows:
                try:        int_nbRows = l_nbRows[i]
                except:     int_nbRows = 0
                try:        int_rowsWhere = l_rowsWhere[i]
                except:     int_rowsWhere = 1
            # FILL THE SHEET
            fStr_fillXls_celByCel(str_path, df_data, str_SheetName, inst_xlApp.xl_lastWsh, int_nbRows, int_rowsWhere)
        
        # Visible and close Excel at the end
        inst_xlApp.Visible = True
        inst_xlApp.CloseWorkbook(True)
        inst_xlApp.QuitXlApp(bl_force = False)
    except:
        try: 		inst_xlApp.Visible = True
        except: 	print('  ERROR in fStr_fillXls_celByCel_plsSheets: xlApp visible did not work')
        try: 		inst_xlApp.CloseWorkbook(True)
        except: 	print('  ERROR in fStr_fillXls_celByCel_plsSheets: Excel workbook could not be closed')
        try: 		inst_xlApp.QuitXlApp(bl_force = False)
        except: 	print('  ERROR: Excel could not be closed')
        print('  ERROR in fStr_fillXls_celByCel_plsSheets: Could not create the PCF: ' + str_path)
        return False
    return str_path

# 1 file out - 1 Dataframe - 1 Sheet
def fStr_fillXls_celByCel(str_path, df_data, str_SheetName = '', xlWs = 0, int_nbRows = 0, int_rowsWhere = 1):
    try:
        # If Sheet is nothing, we must define it
        if xlWs == 0:
            bl_CloseExcel = True
            inst_xlApp = c_win32_xlApp()
            inst_xlApp.FindXlApp(bl_visible = False)
            inst_xlApp.OpenWorkbook(str_path)
            xlWs = inst_xlApp.DefineWorksheet(str_SheetName, 1)
            if not xlWs:        print('  (--) ERROR in fStr_fillXls_celByCel: really could not find the sheet')
        else:   bl_CloseExcel = False
             
        # ------ Insert or delete ROWS ------
        if int_nbRows > 0:
            for i in range(0, int_nbRows):      xlWs.Rows(int_rowsWhere).EntireRow.Insert()
        elif int_nbRows < 0:
            for i in range(0, -int_nbRows):     xlWs.Rows(int_rowsWhere).EntireRow.Delete()                    
        # ------ Fill Cell by Cell  ------
        for i, row in enumerate(df_data.index):
            for j, col in enumerate(df_data.columns):	
                xlWs.Cells(i+1, j+1).Value = str(df_data.iat[i, j])
    except Exception as err:
        print('  ERROR in fl.fStr_fillXls_celByCel: Could not fill  Excel | {}'.format(str(err)))
        print('  - str_path : ', str_path)
        print('  - str_SheetName : ', str_SheetName)
        try:
            print('  - int_nbRows : ', int_nbRows, 'int_rowsWhere: ', int_rowsWhere)
            print('  - i : ', i, 'row: ', row)
            print('  - j : ', j, 'col: ', col)
            print('  - str(df_data.iat[i, j])', str(df_data.iat[i, j]))
            print('  - xlWs.Cells(i+1, j+1).Value', xlWs.Cells(i+1, j+1).Value)
        except:     pass
        return False
    # rustine depending where Function start
    if bl_CloseExcel:
        try:                        inst_xlApp.Visible = True
        except Exception as err:    print('  ERROR in fl.fStr_fillXls_celByCel: xlApp visible did not work | {}'.format(str(err)))
        try:                        inst_xlApp.CloseWorkbook(True)
        except Exception as err:    print('  ERROR in fl.fStr_fillXls_celByCel: Excel workbook could not be closed | {}'.format(str(err)))
        try:                        inst_xlApp.QuitXlApp(bl_force = False)
        except Exception as err:    print('  ERROR: Excel could not be closed | {}'.format(str(err)))
    return str_path

def Act_win32OpenSaveXls(str_path):
    inst_xlApp = c_win32_xlApp()
    inst_xlApp.FindXlApp(bl_visible = True)
    inst_xlApp.OpenWorkbook(str_path)
    inst_xlApp.CloseWorkbook(True)
    return True

def OpenSaveXls_xlWing(str_path):
    inst_xlWings = c_xlApp_xlwings()
    inst_xlWings.FindXlApp(bl_visible = True, bl_screen_updating = False, bl_display_alerts = False)
    inst_xlWings.OpenWorkbook(str_path)
    inst_xlWings.close_Book(bl_saveBeforeClose = True)
    return True


# 1 file out - n Dataframe - n Sheet
def fStr_fillXls_df_xlWgs_sevSh(str_folder,str_FileName,l_dfData,l_SheetName=[],l_nbRows=[],l_rowsWhere=[]):
    try:
        str_path = fStr_BuildPath(str_folder, str_FileName)
        # Open the file (win32)
        inst_xlWings = c_xlApp_xlwings()
        inst_xlWings.FindXlApp(bl_visible = True, bl_screen_updating = False, bl_display_alerts = False)
        inst_xlWings.OpenWorkbook(str_path)
        # Dataframe
        for i in range(len(l_dfData)):
            df_data = l_dfData[i]
            try:        str_SheetName = l_SheetName[i]
            except:     str_SheetName = ''
            #Sheet
            inst_xlWings.DefineWorksheet(str_SheetName, i + 1)
            # ------ Insert or delete ROWS ------
            if l_nbRows:
                try:        int_nbRows = l_nbRows[i]
                except:     int_nbRows = 0
                try:        int_rowsWhere = l_rowsWhere[i]
                except:     int_rowsWhere = 1
            # FILL THE SHEET
            fStr_fillXls_df_xlWgs(str_path, df_data, str_SheetName, inst_xlWings.xl_lastSheet, int_nbRows, int_rowsWhere)
        
        # Close Wk
        inst_xlWings.close_Book(bl_saveBeforeClose = True)
        #inst_xlWings.Quit_xlApp()
    except:
        print('  ERROR in fStr_fillXls_df_xlWgs_sevSh: Could not create the PCF: ' + str_path)
        try: 		
            inst_xlWings.visible = True
            inst_xlWings.screen_updating = True
            inst_xlWings.display_alerts = True
        except: 	print('  ERROR in fStr_fillXls_df_xlWgs_sevSh: visible & Cie did not work')
        return False
    return str_path


# 1 file out - 1 Dataframe - 1 Sheet
def fStr_fillXls_df_xlWgs(str_path, df_data, str_SheetName = '', xl_sheet = 0, int_nbRows = 0, int_rowsWhere = 1):
    try:
        inst_xlWings = c_xlApp_xlwings()
        
        # If Sheet is nothing, we must define it
        if xl_sheet == 0:
            bl_CloseExcel = True    # We do not close the workbook if we need to fill several Sheet
            inst_xlWings.FindXlApp(bl_visible = True, bl_screen_updating = False, bl_display_alerts = False)
            inst_xlWings.OpenWorkbook(str_path)
            xl_sheet = inst_xlWings.DefineWorksheet(str_SheetName, 1)
            if not xl_sheet:        print('  (--) ERROR in fStr_fillXls_celByCel: really could not find the sheet')
        else:   
            bl_CloseExcel = False
            inst_xlWings.xl_lastSheet = xl_sheet
        
        # ------ Insert or delete ROWS ------
        if int_nbRows > 0:
            for i in range(0, int_nbRows):      xl_sheet.range("{0}:{0}".format(str(int_rowsWhere))).api.Insert()
        elif int_nbRows < 0:
            for i in range(0, -int_nbRows):     xl_sheet.range("{0}:{0}".format(str(int_rowsWhere))).api.Delete()
        # ------ Fill DF------
        inst_xlWings.InsertDf_inRange(df_data)
        # Close if only one sheet
        if bl_CloseExcel:
            try:                        inst_xlWings.close_Book(bl_saveBeforeClose = True)
            except Exception as err:    print('  ERROR in fl.fStr_fillXls_df_xlWgs: close_Book | {}'.format(str(err)))
    except Exception as err:
        print('  ERROR in fl.fStr_fillXls_df_xlWgs | {}'.format(str(err)))
        print('  - str_path : ', str_path)
        print('  - str_SheetName : ', str_SheetName)
        print('  - int_nbRows : ', int_nbRows, 'int_rowsWhere: ', int_rowsWhere)
        try: 		
            inst_xlWings.visible = True
            inst_xlWings.screen_updating = True
            inst_xlWings.display_alerts = True
        except: 	print('  ERROR in fStr_fillXls_df_xlWgs: visible & Cie did not work')
        return False
    return str_path






def fTup_GetLastRowCol(xl_sh, int_rowStart = 1, int_colStart = 1):
    int_row = int_rowStart
    int_col = int_colStart
    while xl_sh.Cells(int_row, int_colStart).Value != None:
        int_row += 1
    int_lastRow = int_row - 1
    
    while xl_sh.Cells(int_rowStart, int_col).Value != None:
        int_col += 1
    int_lastCol = int_col - 1
    #    print(xl_sh.Name, 'last_row', int_lastRow, 'last_col', int_lastCol)
    return int_lastRow, int_lastCol


# UK (ASEAN40): SGFABASKETFILETO 
def fDf_readExcelWithPassword(str_path, str_SheetName, str_ExcelPassword):
    d_data = {}
    inst_xlApp = c_win32_xlApp()
    inst_xlApp.FindXlApp(bl_visible = True)
    inst_xlApp.xlApp.DisplayAlerts = False
    inst_xlApp.OpenWorkbook(str_path, str_ExcelPassword)
    
    # Sheet 1: Summary
    xl_sh = inst_xlApp.DefineWorksheet('Summary', 1)
        # Do we need to take only the last row ???????
    int_lastRow, int_lastCol = fTup_GetLastRowCol(xl_sh, 2000, 1)
    rg_content = xl_sh.Range(xl_sh.Cells(7, 1), xl_sh.Cells(int_lastRow, 31)).Value
    df_Summary = pd.DataFrame(list(rg_content))
    # Sheet 2: Equity
    xl_sh = inst_xlApp.DefineWorksheet('Equity', 2)
    int_lastRow, int_lastCol = fTup_GetLastRowCol(xl_sh, 9, 1)
    rg_content = xl_sh.Range(xl_sh.Cells(9, 1), xl_sh.Cells(int_lastRow, int_lastCol)).Value
    df_Equity = pd.DataFrame(list(rg_content))
    # Sheet 8: Div
    xl_sh = inst_xlApp.DefineWorksheet('Div', 8)
    int_lastRow, int_lastCol = fTup_GetLastRowCol(xl_sh, 170, 1)
    rg_content = xl_sh.Range(xl_sh.Cells(5, 1), xl_sh.Cells(int_lastRow, int_lastCol)).Value
    df_Div_Data = pd.DataFrame(list(rg_content))
    # Sheet 9: Est Cash
    xl_sh = inst_xlApp.DefineWorksheet('Est Cash', 9)
    rg_content = xl_sh.Range(xl_sh.Cells(1, 1), xl_sh.Cells(40, 5)).Value
    df_EstCash = pd.DataFrame(list(rg_content))
    # Add columns
    df_Summary = dframe.fDf_Make1stRow_columns(df_Summary)
    df_Equity = dframe.fDf_Make1stRow_columns(df_Equity)
    df_Div_Data = dframe.fDf_Make1stRow_columns(df_Div_Data)
    df_EstCash = dframe.fDf_Make1stRow_columns(df_EstCash)
    # Remove Data
    df_Summary = df_Summary.iloc[-10:].copy()
    df_Summary.reset_index(drop = True, inplace = True)
    # RETURN 
    d_data['Summary'] = df_Summary
    d_data['Equity'] = df_Equity
    d_data['Div'] = df_Div_Data
    d_data['Est Cash'] = df_EstCash
    # CLose
    inst_xlApp.CloseWorkbook(True)
    inst_xlApp.QuitXlApp(bl_force = False)
    
    return d_data
    
    


def fDf_convertToXlsx(str_path, str_SheetName = '', bl_header = None):
    str_pathNew = str_path.replace('.xls', '.xlsx').replace('.XLS', '.xlsx')
    # Test if file exist
    if not fBl_FileExist(str_pathNew):
        try:
            #time.sleep(2)
            inst_xlApp = c_win32_xlApp()
            str_WhereErr = 'Find App'
            inst_xlApp.FindXlApp(bl_gencache_EnsureDispatch = False, bl_visible = True)
            str_WhereErr = 'OpenWorkbook'
            inst_xlApp.OpenWorkbook(str_path)
            str_WhereErr = 'SaveAs'
            inst_xlApp.SaveAs(str_newPath = str_pathNew, int_fileFormat = 51)
            str_WhereErr = 'CloseWorkbook'
            inst_xlApp.CloseWorkbook(False)   
        except Exception as err:
            print('  ERROR fDf_convertToXlsx: {1} - {0} '.format(str(err), str_WhereErr))
            raise            
        print('  (*) Copying XLS file into XLSX: {}'.format(str_pathNew))
    try:
        if str_SheetName == '':     df_data = pd.read_excel(str_pathNew, header = bl_header)
        else:                       df_data = pd.read_excel(str_pathNew, header = bl_header, sheet_name = str_SheetName)
    except:
        print(' ERROR: fDf_convertToXlsx')
        print(' - str_path : ', str_path)
        print(' - str_SheetName : ', str_SheetName)
        print(' - bl_header : ', bl_header)
        raise
    return df_data




#-----------------------------------------------------------------
# FORMAT / STYLE
#-----------------------------------------------------------------
def fStr_StyleIntoExcel(str_path, str_SheetName = '', l_row = [1], str_styleName = 'Header_Perso',
                     bl_bold = True, str_fontColor = '', l_Fill = [], l_border = []):
    # Define EXCEL objects
    xlWb = openpyxl.load_workbook(filename = str_path)
    if str_SheetName == '':     xlWs = xlWb.active
    else:                       xlWs = xlWb.Sheets(str_SheetName)
    
    # Define the Style
    style_header = styl.NamedStyle(name = str_styleName)
    # FONT
    if str_fontColor == 'RED':      style_header.font = styl.Font(bold = bl_bold, color = styl.colors.RED)
    elif str_fontColor == 'BLUE':   style_header.font = styl.Font(bold = bl_bold, color = styl.colors.BLUE)
    elif str_fontColor == 'WHITE':  style_header.font = styl.Font(bold = bl_bold, color = styl.colors.WHITE)
    elif str_fontColor != '':       style_header.font = styl.Font(bold = bl_bold, color = str_fontColor)
    else:                           style_header.font = styl.Font(bold = bl_bold)
    # FILL
    if l_Fill:                      style_header.fill = styl.PatternFill(patternType = l_Fill[0],
                                                                         fill_type = l_Fill[1], 
                                                                         fgColor = l_Fill[2])
    # TO convert the color, please use: https://convertingcolors.com/rgb-color-91_155_213.html
    
    # BORDER
    if l_border: 
        o_border = styl.Side(border_style = l_border[0], color = l_border[1])
        style_header.border = styl.Border(top = o_border, right = o_border, bottom = o_border, left = o_border)
    
    # Save the Style in WK
    try:        xlWb.add_named_style(style_header)
    except:     print(' Information: The Style {} already exists in the workbook'.format(str_styleName))
    
    # Add Celle by Cell
    if 'header' in str_styleName.lower():
        for i_headerRow in l_row:
            header_row = xlWs[i_headerRow]
            for cell in header_row:
                if cell.value != '' and cell.value != None:
                    cell.style = str_styleName
    elif 'table' in str_styleName.lower():
        for i_row in l_row:
            ROW = xlWs[i_row]
            i_nbCol = 0
            for cell in ROW:
                if cell.value != '' and cell.value != None:
                    i_nbCol += 1
            for i_numRow in range(1, 1000):
                ROW = xlWs[i_row + i_numRow]
                str_firstCell = ROW[0].value
                if str_firstCell != '' and str_firstCell != None:
                    for cell in ROW[:i_nbCol]:
                        cell.style = str_styleName
                else:   break
    # SAVE
    xlWb.save(filename = str_path)
    return str_path
 #----------------------------------------


#dte_date = dt.datetime.now().date()
#dd  = dte_date.strftime("%-m")
#print(dd)
#print(type(dd))


# DOCUMENTATION : https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/alignment.html
 
## NIKKO - Basket & Fund
#{
#    'A10:H100':{'font':{'name':'Calibri', 'size':9}},
#    'C11:C100':{'font':{'name':'Calibri', 'size':9},'alignment':{'horizontal':'left'}},
#    'A1:H9':{'font':{'name':'Calibri', 'size':10, 'bold':True}},
#    'B1:B1':{'date':'dd-mmm-yy','font':{'name':'Calibri', 'size':10, 'bold':True}, 
#             'alignment':{'horizontal':'left'}, 
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'B2:B8':{'font':{'name':'Calibri', 'size':10, 'bold':True},
#             'alignment':{'horizontal':'right'}, 
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'A10:H10':{'font':{'size':9,'bold':True,'color':styl.colors.WHITE}, 
#               'alignment':{'vertical':'center', 'wrapText':True}, 
#               'fill':{'patternType':'solid','fill_type':'solid','fgColor': '4F81BD'}},
#    'A3:A3':{'font':{'size':10,'bold':True,'color':styl.colors.WHITE}, 'fill':{'patternType':'solid','fill_type':'solid','fgColor': '4F81BD'}},
#    'A1:A1':{'font':{'size':10,'bold':True,'color':styl.colors.WHITE}, 'fill':{'patternType':'solid','fill_type':'solid','fgColor': '808080'}},
#    'Column_size':{'A':34,'B':24},
#    'Row_size':{10:40},
#    'Table_bord':{'A2:B8':'normBlack', 'B2:B8':'normBlack'},
#    'Table_bord_full':{'A1:B1':'normBlack', 'A2:A2':'normBlack'},
#    'Table_bord_EndDown':{'A10':'normBlack'},
#    'num_format_col':{'E10':'#,##0','F10':'#,##0.00','G10':'#,##0.00','H10':'0.00%'},
#}
#    
### NIKKO - SGX
#{
#    'A1:H100':{'font':{'name':'Arial', 'size':10}},
#    'B2:D16':{'font':{'name':'Arial', 'size':10, 'bold':True}},
#    'B6:B6':{'font':{'name':'Arial', 'size':10, 'bold':True,'color':styl.colors.WHITE},
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': '002060'}},
#    'B4:B4':{'font':{'name':'Arial', 'size':10, 'bold':True,'color':styl.colors.WHITE},
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': '808080'}},
#    'C4:D4':{'date':'dd-mmm-yy','font':{'name':'Arial', 'size':10, 'bold':True},
#              'alignment':{'horizontal':'left'}, 
#              'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'C5:D16':{'font':{'name':'Arial', 'size':10, 'bold':True},
#              'alignment':{'horizontal':'right'}, 
#              'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'B19':{'font':{'name':'Arial', 'size':10},
#           'alignment':{'horizontal':'left'}, 
#           'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'B19:H19':{'font':{'name':'Arial', 'size':10, 'bold':True,'color':styl.colors.WHITE},
#               'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': '808080'}},
#    'Column_size':{'A':4,'B':73,'C':45,'D':14,'E':25,'F':25},
#    'Table_bord':{'B4:D16':'normBlack','B4:B16':'normBlack','C4:D4':'normBlack','B5:B5':'normBlack'},
#    'Table_bord_EndDown_full':{'B19':'normBlack'},
#    'num_format_col':{'D19':'0.00','E19':'0.0000'},
#    'num_format':{'C15:C15':'#,##0.00','C16:C16':'0.00'}
#}
 
## ChinaAMC
#{'A1:K100':{'font':{'name':'Arial', 'size':10}},
#    'B2:C20':{'font':{'name':'Arial', 'size':10, 'bold':True}},
#    'B8:B8':{'font':{'name':'Arial', 'size':10, 'bold':True,'color':styl.colors.WHITE},
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': '002060'}},
#    'B5:B6':{'font':{'name':'Arial', 'size':10, 'bold':True,'color':styl.colors.WHITE},
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': '808080'}},
#    'B22:K22':{'font':{'name':'Arial', 'size':10, 'bold':True,'color':styl.colors.WHITE},
#               'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': '808080'}},
#    'C5:C6':{'date':'m/d/yyyy',
#             'font':{'name':'Arial', 'size':10, 'bold':True},
#             'alignment':{'horizontal':'left'}, 
#             'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},             
#    'C8:C20':{'font':{'name':'Arial', 'size':10, 'bold':True},'alignment':{'horizontal':'right'},
#              'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'B23':{'font':{'name':'Arial', 'size':10},'alignment':{'horizontal':'left'},
#           'fill':{'patternType':'solid', 'fill_type':'solid', 'fgColor': 'F2F2F2'}},
#    'Column_size':{'A':4,'B':75,'C':70,'D':12,'G':21,'K':25},
#    'Table_bord':{'B9:C20':'normBlack','C9:C20':'normBlack'},
#    'Table_bord_full':{'B5:C6':'normBlack','B8:C8':'normBlack'},
#    'Table_bord_EndDown_full':{'B22':'normBlack'},
#    'num_format':{'C14:C14':'#,##0.0000','C15:C17':'#,##0','C18:C19':'#,##0.00','C20:C20':'#,##0'},
#    'num_format_col':{'E22':'0.00','F22':'0.00','G22':'0.00000000000000','K22':'0.00000000000000'}
#}

## WisdomTree
#{'A1:U1':{'font':{'bold':True,'color':styl.colors.WHITE}, 'fill':{'patternType':'solid','fill_type':'solid','fgColor': '4F81BD'}}, 
#    'A5:G5':{'font':{'bold':True,'color':styl.colors.WHITE}, 'fill':{'patternType':'solid','fill_type':'solid','fgColor': '4F81BD'}}, 
#    'A10:T10':{'font':{'bold':True,'color':styl.colors.WHITE}, 'fill':{'patternType':'solid','fill_type':'solid','fgColor': '4F81BD'}},
#    'Table_bord_EndDown_full':{'A1':'WT_blue', 'A5':'WT_blue', 'A10':'WT_blue'}     
#}





def Act_StyleIntoExcel(str_path, str_format = '', str_SheetName = ''):
    # EVAL
    try:
        if str_format == '':    return True
        d_format = eval(str_format)
    except Exception as err:
        print(' ERROR Act_StyleIntoExcel - EVAL')
        print(' - Err :', err)
        return False
    # Define EXCEL objects
    try:
        xlWb = openpyxl_Excel.load_workbook(filename = str_path)
        if str_SheetName == '':     xlWs = xlWb.active
        else:                       xlWs = xlWb[str_SheetName]
    except Exception as err:
        print(' ERROR Act_StyleIntoExcel, could not define the sheet')
        print(' - Err :', err, '\n')
        return False
    # Launch the different process included into the dico
    try:
        for str_area, d_formatValue in d_format.items():
            if 'column_size' in str_area.lower():
                Act_resizeRowColumn(xlWs, 'column', d_formatValue)
            elif 'row_size' in str_area.lower():
                Act_resizeRowColumn(xlWs, 'row', d_formatValue)
            elif 'Table_bord'.lower() in str_area.lower():
                str_keyType = str_area          # Area is now Table_bord / Table_bord_endDown....
                if '_full' in str_keyType.lower():          bl_full = True
                else:                                       bl_full = False
                d_border = d_format[str_keyType]
                for str_areaBorder, str_borderName in d_border.items():
                    if '_EndDown'.lower() in str_keyType.lower():
                        rg_toSelect = fRg_SelectRangeToApplyFormat(xlWs, str_areaBorder, bl_includeHeader = True)
                    else:
                        rg_toSelect = xlWs[str_areaBorder]
                    Act_loopBorder(str_keyType, rg_toSelect, str_borderName, bl_full)
            elif 'num_format' in str_area.lower():
                str_keyType = str_area          # Area is now num_format
                d_colParam = d_format[str_keyType]
                for str_areaFormat, str_format in d_colParam.items():
                    if '_col' in str_keyType.lower():
                        rg_toSelect = fRg_SelectRangeToApplyFormat(xlWs, str_areaFormat, bl_includeHeader = False, bl_column = True)
                    else:
                        rg_toSelect = xlWs[str_areaFormat]
                    Act_loopFormat(rg_toSelect, str_format, 'num_format')
            else:
                str_styleName = fStr_defineStyle(xlWb, d_formatValue)
                # Define Aera if we just put one cell as input
                if ':' in str_area:     rg_toSelect = xlWs[str_area]
                else:                   rg_toSelect = fRg_SelectRangeToApplyFormat(xlWs, str_area, bl_includeHeader = True)
                Act_loopFormat(rg_toSelect, str_styleName)
    except Exception as err:
        print(' ERROR Act_StyleIntoExcel: Loop on Area for Style')
        print(' - Err :', err, '\n')
    # SAVE
    try:    xlWb.save(filename = str_path)
    except Exception as err:
        print(' ERROR Act_StyleIntoExcel, xlWb.save')
        print(' - Err :', err)
        return False
    return True

#-----------------------------------------------------
def fStr_defineStyle(xlWb, d_formatValue):
    # Define and add a Style format depending on a name dev created (NikkoHeader_Blue)
    try:
        # Define the Style - Name
        # {'font':{'name':'Calibri', 'size':9}}
        l_styleName = list(d_formatValue.keys())
        for dic in list(d_formatValue.values()):
            if isinstance(dic, dict):
                l_styleName.extend(list(dic.keys()))
                l_styleName.extend(list(dic.values()))
        str_styleName = '_'.join([str(x) for x in l_styleName])
        str_styleName = str_styleName.replace(' ', '')
        # Format Date
        if 'date' in d_formatValue:
            str_formatDate = d_formatValue['date']
            o_style = styl.NamedStyle(name = str_styleName, number_format = str_formatDate)
        else:
            o_style = styl.NamedStyle(name = str_styleName)
        
        # Conditional
        if 'font' in d_formatValue:
            d_font = d_formatValue['font']
            o_style.font = styl.Font(**d_font)
        if 'fill' in d_formatValue:
            d_fill = d_formatValue['fill']
            o_style.fill = styl.PatternFill(**d_fill)
        if 'alignment' in d_formatValue:
            d_align = d_formatValue['alignment']
            o_style.alignment = styl.Alignment(**d_align)
            
    except Exception as err:
        print(' ERROR fStr_defineStyle: Loop on Area for Style')
        print(' - Err :', err)
        print(' - d_formatValue :', d_formatValue)
        print(' - str_styleName :', str_styleName)
        print(' - o_style :', o_style)
        raise
    # Save the Style in WK
    try:        xlWb.add_named_style(o_style)
    except:     print(' Information: Style already exists in the workbook: {}'.format(str_styleName))
    return str_styleName

#-----------------------------------------------------
def Act_loopFormat(l_rows, str_styleName, str_type = ''):
    # Loop Cell by Cell to apply a format
    try:
        for row in l_rows:
            for cell in row:
                if 'num_format' in str_type.lower():
                    cell.number_format = str_styleName
                else:
                    cell.style = str_styleName
    except Exception as err:
        print(' ERROR Act_loopFormat: Loop on Area for Style')
        print(' - Param :', str_styleName)
        print(' - Err :', err)
        try:
            print(' - row :', row)
            print(' - cell :', cell)
        except: pass
        raise

#-----------------------------------------------------
def Act_resizeRowColumn(xlWs, str_type, d_formatValue):
    try:
        if 'col' in str_type:
            for col in d_formatValue:
                col_dimension = d_formatValue[col]
                if isinstance(col_dimension, int):
                    xlWs.column_dimensions[col].width = col_dimension
                else:   print(' ERROR in Act_resizeRowColumn - Column_size need to be an integer')
        elif 'row' in str_type:
            for row in d_formatValue:
                row_dimension = d_formatValue[row]
                if isinstance(row_dimension, int):
                    xlWs.row_dimensions[row].height = row_dimension
                else:   print(' ERROR in Act_resizeRowColumn - row_dimension need to be an integer')
    except Exception as err:
        print(' ERROR Act_resizeRowColumn')
        print(' - Err :', err)
        raise

#-----------------------------------------------------
def Act_loopBorder(str_type, rg_toSelect, str_borderName, bl_full = False):
    try:
        if str_borderName == 'normBlack':
            o_border = styl.Side(border_style = 'thin')
        elif str_borderName == 'WT_blue':
            o_border = styl.Side(border_style = 'thin', color = '4A7FB0')
        else:       print('Please define the Border in Act_loopBorder || {}'.format(str_borderName))
        
        #===========================
        # Full Array => *Ignore all below condition
        if bl_full:
            for row in rg_toSelect:
                for cell in row:
                    cell.border =   styl.Border(top = o_border, bottom = o_border, left = o_border, right = o_border)
            return True
        
        #===========================
        # Get the characteristics of the Array
        if rg_toSelect[0] == rg_toSelect[-1]:
            bl_uniqueRow = True
            if rg_toSelect[0][0] == rg_toSelect[0][-1]:
                bl_uniqueCell = True
            else:   bl_uniqueCell = False
        else:
            bl_uniqueCell = False
            bl_uniqueRow = False
            if rg_toSelect[0][0] == rg_toSelect[0][-1]:
                bl_uniqueCol = True
            else:   bl_uniqueCol = False
        #===========================
        # I. One Cell
        if bl_uniqueCell:
            cell = rg_toSelect[0][0]
            cell.border =   styl.Border(top = o_border, bottom = o_border, left = o_border, right = o_border)
            return True
        #===========================
        # II. One Row
        if bl_uniqueRow:
            row_unique = rg_toSelect[0]
            for cell in row_unique:
                # left cell
                if cell == row_unique[0]:
                    cell.border =   styl.Border(top = o_border, bottom = o_border, left = o_border)
                # right cell
                elif cell == row_unique[-1]:
                    cell.border =   styl.Border(top = o_border, bottom = o_border, right = o_border)
                else: cell.border = styl.Border(top = o_border, bottom = o_border)
            return True
        #===========================
        # II. One column
        if bl_uniqueCol:
            for row in rg_toSelect:
                cell = row[0]
                # Top cell
                if cell == rg_toSelect[0][0]:
                    cell.border =   styl.Border(left = o_border, right = o_border, top = o_border)
                # bottom cell
                elif cell == rg_toSelect[-1][0]:
                    cell.border =   styl.Border(left = o_border, right = o_border, bottom = o_border)
                else: cell.border = styl.Border(left = o_border, right = o_border)
            return True
        #===========================
        # III. Proper Array
        # III.a. Loop for Left and right
        for row in rg_toSelect:
            cell_left = row[0]
            cell_right = row[-1]
            cell_left.border =      styl.Border(left = o_border)
            cell_right.border =     styl.Border(right = o_border)
        # III.b. Loop for Top 
        row_top = rg_toSelect[0]
        for cell in row_top:
            # Top-left cell
            if cell == rg_toSelect[0][0]:
                cell.border =   styl.Border(top = o_border, left = o_border)
            # Top-right cell
            elif cell == rg_toSelect[0][-1]:
                cell.border =   styl.Border(top = o_border, right = o_border)
            else: 
                cell.border =   styl.Border(top = o_border)
        # III.c. Loop for bottom
        row_bottom = rg_toSelect[-1]
        for cell in row_bottom:
            # bottom-left cell
            if cell == rg_toSelect[-1][0]:
                cell.border =   styl.Border(bottom = o_border, left = o_border)
            # bottom-right cell
            elif cell == rg_toSelect[-1][-1]:
                cell.border =   styl.Border(bottom = o_border, right = o_border)
            else: 
                cell.border =   styl.Border(bottom = o_border)
    except Exception as err:
        print(' ERROR Act_loopBorder')
        print(' - Err :', err)
        print(' - Param :', str_type, str_borderName)
        print(' - rg_toSelect :', rg_toSelect)
        try:    print(' - row :', row)
        except: pass
        try:    print(' - cell :', cell)
        except: pass
        raise


#-----------------------------------------------------
def fRg_SelectRangeToApplyFormat(xlWs, str_cell, bl_includeHeader = True, bl_column = False):
    try:
        d_colNumber = {0:'A', 1:'A', 2:'B', 3:'C', 4:'D', 5:'E', 6:'F', 7:'G', 8:'H', 9:'I', 10:'J', 11:'K', 
                       12:'L', 13:'M', 14:'N', 15:'O', 16:'P', 17:'Q', 18:'R', 19:'S', 20:'T', 21:'U', 
                       22:'V', 23:'W', 24:'X', 25:'Y', 26:'Z', 27:'AA', 28:'AB', 29:'AC', 30:'AD', 
                       31:'AE', 32:'AF', 33:'AG', 34:'AH', 35:'AI', 36:'AJ', 37:'AK', 38:'AL', 39:'AM'}
        # CAREFULL (still just for one letter in the column)
        str_colStart = str_cell[0]
        # find The column number
        d_NumberCol = {value:key for key,value in d_colNumber.items() if key > 0}
        int_colBase = d_NumberCol[str_colStart]
        # Find the Row object
        int_rowHeader = int(str_cell[1:])
        row_header = xlWs[int_rowHeader]
        #print(str_colStart, int_rowHeader, row_header)
        #---------------------
        # Get the Max row
        i_rowNumFin = int_rowHeader
        for i_numRow in range(1, 10000):
            ROW = xlWs[int_rowHeader + i_numRow]
            if ROW[int_colBase-1].value == '' or ROW[int_colBase-1].value == None:  break
            i_rowNumFin += 1
        # include Header or not
        if not bl_includeHeader:    int_rowHeader += 1
        #---------------------
        # Just on the column
        if bl_column:
            str_area = "{}{}:{}{}".format(str_colStart, int_rowHeader, str_colStart, i_rowNumFin)
        else:
            #---------------------
            # Get the Max column
            i_colIter = 0
            i_colNumFin = int_colBase - 1
            for cell in row_header:
                if i_colIter < i_colNumFin:
                    # If we start at column B we dont want to take column A into account
                    i_colIter += 1
                else:
                    if cell.value == '' or cell.value == None:  break
                    i_colNumFin += 1
                    i_colIter += 1
            # Final
            str_area = "{}{}:{}{}".format(str_colStart, int_rowHeader, d_colNumber[i_colNumFin], i_rowNumFin)
        #print(i_colNumFin, str_area)
        #---------------------
        # Define Range
        rg_toSelect = xlWs[str_area]
        #---------------------
    except Exception as err:
        print(' ERROR fRg_SelectRangeToApplyFormat')
        print(' - Err :', err)
        print(' - Param :', str_cell, bl_includeHeader)
        print(' - i_colNumFin :', i_colNumFin)
        print(' - i_rowNumFin :', i_rowNumFin)
        print(' - str_area :', str_area)
        raise
    return rg_toSelect



















#-----------------------------------------------------------------
# ZIP
#-----------------------------------------------------------------
def ZipExtractFile(str_ZipPath, str_pathDest = '', str_FileName = '', bl_extractAll = False, str_zipPassword = ''):
    try:
        with ZipFile(str_ZipPath, 'r') as zipObj:
            if str_zipPassword != '':
                zipObj.setpassword(pwd = bytes(str_zipPassword, 'utf-8'))
                bl_extractAll = True
                #                zipObj.extractall(pwd = bytes(str_zipPassword, 'utf-8'))                
            if bl_extractAll:
                # Extract all the file
                if str_pathDest == '':      zipObj.extractall()
                else:                       zipObj.extractall(str_pathDest)
                time.sleep(5)
            else:
                # Get a list of all archived file names from the zip
                l_fileInZip = zipObj.namelist()
                # Get The name of the list  matching
                l_fileInZip_file = fL_GetFileList_withinModel(l_fileInZip, str_FileName)
                # Extract the file
                for file in l_fileInZip_file:
                    zipObj.extract(file, str_pathDest)
    except Exception as err:
        print(' ERROR: ZipExtractFile || {}'.format(str(err)))
        if bl_extractAll:
            print(' - str_ZipPath : ', str_ZipPath)
            print(' - str_pathDest : ', str_pathDest)
            print(' - str_zipPassword : ', str_zipPassword)
            raise
        else:
            print(' - Failed to download the file : ', str_FileName)
            if l_fileInZip:
                print(' - File List in the Zip : ', l_fileInZip)
            print(' (**) Trying to extract all files...')
            ZipExtractFile(str_ZipPath, str_pathDest, '', True, str_zipPassword)
    return True

#ZipExtractFile(r'C:\Users\laurent.tupin\Documents\5_Python\Py_Package\Brouillon\cccc.zip',
#               r'C:\Users\laurent.tupin\Documents\5_Python\Py_Package\Brouillon',
#               'HKGRFMHKSSET_ST-FM-TX-001-01_20191127_1_{XXXXXX}.XLS')
  
    




#------------------------------------------------------------------------------
#--------- To Archives / Prod --------------
#------------------------------------------------------------------------------   
def fL_GetListSubFolder_except(str_folder, str_folderExcept = ''):
   if str_folderExcept != '':       return [x[0] for x in os.walk(str_folder) if x[0][:9] != str_folderExcept]
   else:                            return [x[0] for x in os.walk(str_folder)]
   

def fL_GetListDirFileInFolders(l_subDir, l_typeFile = []):
    listTuple_PathFic = []
    if l_typeFile:
        for Dir in l_subDir:
            list_fic = [(Dir, fic) for fic in fList_FileInDir(Dir) if fic[-3:].lower() in l_typeFile or fic[-4:].lower() in l_typeFile]
            if list_fic: listTuple_PathFic += list_fic
    else:
        for Dir in l_subDir:
            list_fic = [(Dir, fic) for fic in fList_FileInDir(Dir)]
            if list_fic: listTuple_PathFic += list_fic
    return listTuple_PathFic




def Act_CopyUpdateFiles_specialBackUp(l_FolderFic_from, str_DestFolder, dte_after = 10, str_removeInDestFolder = ''):
    str_message = ''
    
    # Date limite
    if type(dte_after) == int:
        dte_after = dt.datetime.now() - dt.timedelta(dte_after)
    
    # Get list of Origin Files
    l_pathOrigin = [os.path.join(str_folder, str_file) for (str_folder, str_file) in l_FolderFic_from]
    l_pathOrigin_wLimit = fL_KeepFiles_wTimeLimit(l_pathOrigin, dte_after)
    
    # Create the Folder Destination
    try:
        l_folderDest = [fStr_GetFolderFromPath(str_path).replace('.', str_DestFolder, 1).replace(str_removeInDestFolder, 
                        '') for str_path in l_pathOrigin_wLimit]
        l_folderDest_unique = list(set(l_folderDest))
        for folder in l_folderDest_unique:
            fBl_createDir(folder)
    except Exception as err:
        str_message += fStr_Message(' ERROR in Act_CopyUpdateFiles_specialBackUp: could not crete folder in Dest ||| {}'.format(str(err)))
        return str_message
        
    # Get the Destination Path
    l_pathDest = [str_path.replace('.', str_DestFolder, 1).replace(str_removeInDestFolder, '') for str_path in l_pathOrigin_wLimit]
    #    l_pathDest = [os.path.join(str_folder.replace('.', str_DestFolder).replace(str_removeInDestFolder, ''), str_file) for 
    #                  (str_folder, str_file) in l_FolderFic_from]
    
    # Loop on File to copy / update them
    for i_it in range(len(l_pathOrigin_wLimit)):
        str_pathOrigin = l_pathOrigin_wLimit[i_it]
        str_pathDest = l_pathDest[i_it]
        # If File DOES NOT Exists
        if not fBl_FileExist(str_pathDest):
            str_message += fStr_Message('COPY...  Origin: {} ||| Dest: {}'.format(str_pathOrigin, fStr_GetFolderFromPath(str_pathDest)))
            try:        shutil.copy(str_pathOrigin, str_pathDest)
            except:     str_message += fStr_Message(' (--) ERROR: file could not be Copied !!!') 
        else:
            # Compare Date (Update only if CLoud is more recent)
            dte_lastmod = fDte_GetModificationDate(str_pathOrigin)
            dte_lastmod_dest = fDte_GetModificationDate(str_pathDest)
            if dte_lastmod > dte_lastmod_dest:
                str_message += fStr_Message('UPDATE...  Origin: {} ||| Dest: {}'.format(str_pathOrigin, fStr_GetFolderFromPath(str_pathDest)))
                #---ARCHIVES--------------
                if r'\Archive' in str_DestFolder[-10:]:
                    str_dateTime =  str(dte_lastmod_dest.strftime('%Y%m%d'))
                    shutil.copyfile(str_pathDest, fStr_GetFolderFromPath(str_pathDest) + '\\' + str_dateTime + '_' + fStr_GetFileFromPath(str_pathDest))
                #------------------------
                try:        shutil.copy(str_pathOrigin, str_pathDest)
                except:     str_message += fStr_Message(' (--) ERROR: file could not be Updated !!!') 
    str_message += fStr_Message(' ... End CopyPaste Process !!!')
    return str_message


def Act_CopyUpdateFiles(l_PathFic_from, l_PathFic_dest, str_DestFolder = '', str_removeInDestFolder = ''):
    Act_CopyUpdateFiles_specialBackUp(l_PathFic_from, str_DestFolder, dte_after = 900, str_removeInDestFolder = str_removeInDestFolder)
    
    #    if str_DestFolder == '':
    #        print('Fill the 3rd argument on the function: Act_CopyUpdateFiles')
    #        return False
    #    # Loop on File to copy / update them
    #    for t_file in l_PathFic_from:
    #        str_path = t_file[0]
    #        str_file = t_file[1]
    #        str_path_Dest = str_path.replace('.', str_DestFolder)
    #        str_path_Dest = str_path_Dest.replace(str_removeInDestFolder, '')
    #        
    #        # If file is new --> Copy
    #        if (str_path_Dest, str_file) not in l_PathFic_dest:
    #            print('COPY NEW...   ', 'Folder Origin:   ' + str_path, ' ||| Folder Dest:   ' + str_path_Dest, ' ||| File:   ' + str_file)
    #            fBl_createDir(str_path_Dest)
    #            shutil.copyfile(str_path + '\\' + str_file, str_path_Dest + '\\' + str_file)
    #        else:
    #            dte_lastmod = fDte_GetModificationDate(str_path + '\\' + str_file)
    #            dte_lastmod_dest = fDte_GetModificationDate(str_path_Dest + '\\' + str_file)
    #            # Compare Date
    #            if dte_lastmod > dte_lastmod_dest:
    #                print('COPY UPDATE...', 'Folder Origin:   ' + str_path, ' ||| Folder Dest:   ' + str_path_Dest, ' ||| File:   ' + str_file)
    #                if '.\\Archive' in str_DestFolder:
    #                    str_dateTime =  str(dte_lastmod_dest.strftime('%Y%m%d'))
    #                    shutil.copyfile(str_path_Dest + '\\' + str_file, str_path_Dest + '\\' + str_dateTime + '_' + str_file)
    #                shutil.copyfile(str_path + '\\' + str_file, str_path_Dest + '\\' + str_file)
    return True


def fStr_CopPasteFolder(str_folderOrigin, str_folderTarget, dte_after = 10, l_typeFile = [], str_folderExcept = ''):
    # Date limite
    if type(dte_after) == int:
        dte_after = dt.datetime.now() - dt.timedelta(dte_after)

    # Environment of work
    dir_current = os.getcwd()
    os.chdir(str_folderOrigin)
    # Get all the sub Dir in the folder -- Except the folder (if empty, no exception)
    l_SubDir_Origin = fL_GetListSubFolder_except('.', str_folderExcept)
    # Get Tuples in List (Path, File Python)
    l_PathFic = fL_GetListDirFileInFolders(l_SubDir_Origin, l_typeFile)
    # Copy / Update files from a list of tuple to another
    str_message = Act_CopyUpdateFiles_specialBackUp(l_PathFic, str_folderTarget, dte_after)
    # Fin !!
    os.chdir(dir_current)
    return str_message


def Act_CopPasteFolder_severalTry(str_folderDest, l_pathOrigin, dte_after = 10, l_typeFile = [], str_folderExcept = ''):
    for pathOrigin in l_pathOrigin:
        try:
            fStr_CopPasteFolder(pathOrigin, str_folderDest, dte_after = dte_after, l_typeFile = l_typeFile,str_folderExcept = str_folderExcept)
            return True
        except: pass
    return False






#------------------------------------------------------------------------------
# DEPRECATED
#------------------------------------------------------------------------------
    
